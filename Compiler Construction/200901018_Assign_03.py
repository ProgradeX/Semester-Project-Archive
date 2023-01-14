import xml.etree.ElementTree as ElemT
import openpyxl


def main():
    # Parse the XML file
    tree = ElemT.parse('compiler.xml')

    # open Excel worksheet
    workbook = openpyxl.Workbook()
    sheet = workbook.active

    # Get the root element
    root = tree.getroot()

    # Assigning Column Names
    sheet.cell(1, 1).value = 'Book ID'
    sheet.cell(1, 2).value = 'Author Name'
    sheet.cell(1, 3).value = 'Title'
    sheet.cell(1, 4).value = 'Genre'
    sheet.cell(1, 5).value = 'Price'
    sheet.cell(1, 6).value = 'Publish date'
    sheet.cell(1, 7).value = 'Description'

    # Iterate over the children of the root element
    for i, child in enumerate(root):
        # Print the tag and text for each child element
        print(child.tag, child.attrib['id'])
        sheet.cell(row=i + 2, column=1).value = child.attrib['id']

        for j, subChild in enumerate(child):
            # Print the tag and text for each subChild element
            print(f"  {subChild.tag}: {subChild.text}")
            sheet.cell(row=i + 2, column=j+2).value = subChild.text

    workbook.save('200901018_Assign_03.xlsx')


if __name__ == '__main__':
    main()
